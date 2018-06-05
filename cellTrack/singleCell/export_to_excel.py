import openpyxl, re
import matplotlib.pyplot as plt
import directory
from glob import glob
from scipy.signal import savgol_filter


file_directory = directory.fdirectory

book = openpyxl.Workbook()
sheet = book.active
sheet.title = 'Cell Coordinates'

text_mask = file_directory + '*' + '_coordinates.txt'
text_names = glob(text_mask)

c = 1
for fn in text_names:
    raw_data = open(fn, 'r')
    serial = re.findall("\d+", fn)[0]
    data = list(raw_data)

    sheet.cell(row=1, column=c).value = 'cell ' + serial
    # sheet.cell(row=2, column=c).value = data[-2]
    y = []
    x = range(0, len(data[:-2]))
    r = 2
    for index, point in enumerate(data[:-2]):
        zeroed = round(float(point) - float(data[0]), 3)
        y.append(zeroed)
        sheet.cell(row=r + index, column=c).value = zeroed

    yhat = savgol_filter(y, 51, 3)
    plt.plot(x,y)
    plt.plot(x, yhat, color='red')

    plt.xlabel('time (frames)')
    plt.ylabel('position (pixels)')
    plt.title(serial)

    plt.show()
    c += 1

# print('Filename: %s successfully loaded and recorded!' % fn)

book.save(file_directory + '/cell_coodinates.xlsx')

